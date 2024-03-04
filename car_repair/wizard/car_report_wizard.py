from odoo import fields, models, api
from odoo.exceptions import UserError

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64

import base64
from io import BytesIO


class CarReportWizard(models.TransientModel):

    _name = 'car.report.wizard'
    _description = "Car Report"

    file_name = fields.Binary(string="Report")
    datas_fname = fields.Char(string="Filename")
    start_date = fields.Date(string="Start Date")
    end_date = fields.Date(string="End Date")
    detail_ids = fields.Many2many(
        'detail.model', string="records id", compute='_compute_filtered_patient_ids')

    @api.depends('start_date', 'end_date')
    def _compute_filtered_patient_ids(self):
        for record in self:
            domain = [('start_date', '>=', record.start_date),
                      ('end_date', '<=', record.end_date)]

            record.detail_ids = self.env['detail.model'].search(domain)

    def print_pdf(self):
        if self.start_date and self.end_date and self.end_date < self.start_date:
            raise UserError("End date cannot be before start date.")

        return self.env.ref('car_repair.action_report_car_repair_filter').report_action(self)

    def print_xlsx(self):
        if self.start_date and self.end_date and self.end_date < self.start_date:
            raise UserError("End date cannot be before start date.")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Car Report"

        # Define border style
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
    
        # Merge cells for Car Report heading
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

        # Set value for the merged cell (Car Report heading)
        car_report_cell = worksheet.cell(row=1, column=1)
        car_report_cell.value = "Car Report"
        car_report_cell.alignment = Alignment(horizontal='center')
        car_report_cell.font = Font(bold=True, color="000000")
        car_report_cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  # Green background color
        car_report_cell.border = thin_border  # Apply border

        # Merge cells for Start Date header
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)

        # Set value for the merged cell (Start Date)
        start_date_cell = worksheet.cell(row=3, column=1)
        start_date_cell.value = "Start Date"
        start_date_cell.alignment = Alignment(horizontal='center')
        start_date_cell.font = Font(bold=True, color="000000")
        start_date_cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  # Green background color
        start_date_cell.border = thin_border  # Apply border

        # Merge cells for End Date header
        worksheet.merge_cells(start_row=3, start_column=7, end_row=3, end_column=9)

        # Set value for the merged cell (End Date)
        end_date_cell = worksheet.cell(row=3, column=7)
        end_date_cell.value = "End Date"
        end_date_cell.alignment = Alignment(horizontal='center')
        end_date_cell.font = Font(bold=True, color="000000")
        end_date_cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  # Green background color
        end_date_cell.border = thin_border  # Apply border

        # Set values for start date and end date
        start_date_value_cell = worksheet.cell(row=4, column=2)
        start_date_value_cell.value = self.start_date.strftime("%d-%m-%Y")
        start_date_value_cell.alignment = Alignment(horizontal='center')

        end_date_value_cell = worksheet.cell(row=4, column=8)
        end_date_value_cell.value = self.end_date.strftime("%d-%m-%Y")
        end_date_value_cell.alignment = Alignment(horizontal='center')

        # Define column headers and their widths for main data table
        main_headers = [("User", 15), ("Email", 25), ("Car Name", 12), ("License Plate", 15), ("Car Model", 14),
                        ("Oil Charge", 15), ("Washing Charge", 17), ("Service Charge", 17), ("Total", 13)]

        # Apply styles to header cells for main data table
        for header, width in main_headers:
            cell = worksheet.cell(row=5, column=main_headers.index((header, width)) + 1)
            cell.value = header
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")  # Green background color
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align text
            cell.border = thin_border  # Apply border
            worksheet.column_dimensions[get_column_letter(cell.column)].width = width

        # Loop through detail_ids and populate main data table
        row_index = 6
        for record in self.detail_ids:
            total_cost = 0
            for car in record.car_info_ids:
                row_data = [
                    record.client_name.name,
                    record.email,
                    car.car_name,
                    car.license_plate,
                    car.car_model,
                    car.car_oil_price,
                    car.car_washing_charge,
                    car.car_service_charge,
                    car.total
                ]
                # Append row data
                for col_index, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_index, column=col_index)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')  # Center align text
                    cell.border = thin_border  # Apply border
                total_cost += car.total or 0
                row_index += 1

            # Add a row for the total cost of this detail record
            total_row = [""] * 7 + ["Total Cost", total_cost]
            for col_index, value in enumerate(total_row, start=1):
                cell = worksheet.cell(row=row_index, column=col_index)
                cell.value = value
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')  # Center align text
                cell.border = thin_border  # Apply border
            row_index += 2  # Add extra row for spacing after total row

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        file_name = "Car_Report.xlsx"
        self.file_name = base64.b64encode(output.read())
        self.datas_fname = file_name

        return {
            'type': 'ir.actions.act_url',
            'url': "web/content/?model=%s&id=%s&filename_field=%s&field=file_name&download=true&filename=%s" % (
                self._name, self.id, 'datas_fname', file_name),
            'target': 'self',
        }
